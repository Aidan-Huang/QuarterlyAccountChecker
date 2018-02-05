#!/usr/bin/env python
# encoding: utf-8
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

# 去除冗余信息的正则表达式
re1 = r'\/.+?(?=[\s;]|$)'

# 向上找第一个非空单元格
def find_the_first_not_none_value_upward(ws, row, column):
	for i in range(row - 1, 1, -1):
		if ws[column + str(i)].value is not None:
			return ws[column + str(i)]

# 更改电信部门名称为公司部门名称
def change_teleDept_to_idealDept(teleDept):
	deptDict = {
		"党群工作办公室/企业文化处/纪检监察室": "党群工作办公室",
		"电信增值应用软件部": "增值应用业务部",
		"公司领导": "公司总部",
		"电信支撑软件部": "电信支撑软件部",
		"计划财务处": "计划财务部",
		"技术部": "技术质量管理部",
		"人力资源库": "人力资源部"
	}

	if teleDept in deptDict:
		return deptDict[teleDept]
	else:
		return teleDept




# 电信公司OA对账文件
FILE_NAME = '17Q4.xlsx'

# 核对账期 4位年份2位季度
CHECK_TIME = '201704'

# 核对结果表单位置
RESULT_SHEET_NUMBER = 2

# 核对账号类型
ACCOUNT_TYPE = '电信OA账号'

# 加载文件
wb = load_workbook(filename = FILE_NAME)

# 若对账结果表单则先删除
try:
    ws = wb[CHECK_TIME]
    wb.remove(ws)
except KeyError as e:
    print ('No sheet test:' + e.__str__())

# 创建对账结果表单
wb.create_sheet(CHECK_TIME)
wsc = wb[CHECK_TIME]

# 写入标题行
# 序号	部门	姓名	账号	账号类型	对账时间
wsc["A1"] = "序号"
wsc["B1"] = "部门"
wsc["C1"] = "姓名"
wsc["D1"] = "账号"
wsc["E1"] = "账号类型"
wsc["F1"] = "对账时间"

# 标题行黑体
wsc["A1"].font = Font(bold=True)
wsc["B1"].font = Font(bold=True)
wsc["C1"].font = Font(bold=True)
wsc["D1"].font = Font(bold=True)
wsc["E1"].font = Font(bold=True)
wsc["F1"].font = Font(bold=True)

# 当前行号
currentRow = 2

ws = wb['部门信息']

# 在B列找“成员”内容的单元格
# 在A列找相应的部门
# 在C列具体成员删除冗余字符串
# 拷贝 “部门”，“成员”到临时表单
for row in ws.rows:
	for cell in row:
		if cell.value == '成员':
			if cell.column == 'B':
				# print(ws['A' + str(cell.row - 1)].value)
				departName = find_the_first_not_none_value_upward(ws, cell.row, 'A').value
				# print (departName)
				usersStr = ws['C' + str(cell.row )].value
				usersStr = re.sub(re1, '', str(usersStr))
				usersStr= re.sub('admin_lxgs;', '', usersStr)
				usersStr= re.sub(r'\d', '', usersStr)
				usersStr= re.sub('_', '', usersStr)
				if usersStr != "None" and departName not in ["工会", "团委", "财务处", "党委"]:

					departName = change_teleDept_to_idealDept(departName)

					users = usersStr.split(';')
					for user in users:
						rowStr = str(currentRow)

						wsc["A" + rowStr] = str(currentRow - 1)
						wsc["B" + rowStr] = departName
						wsc["C" + rowStr] = user
						wsc["D" + rowStr] = ''
						wsc["E" + rowStr] = ACCOUNT_TYPE
						wsc["F" + rowStr] = CHECK_TIME

						currentRow += 1

wb.active = RESULT_SHEET_NUMBER
wb.save(FILE_NAME)

print ("Total " + str(currentRow - 2) + " employees.")