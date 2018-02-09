#!/usr/bin/env python
# encoding: utf-8
from openpyxl import load_workbook
from openpyxl.styles import Font,colors

# 找第一个非空单元格
def find_row_count_at_column(sheet, row, column):
    count = 0
    for i in range(row, 3000):
        if sheet.cell(row=i, column=column).value is None:
            return count
        else:
            count += 1

# 数列写入Excel行
def list_to_row(data, sheet, row, column):
    for i in range(0, list(data).__len__()):
        sheet.cell(row=row, column=column + i).value = data[i]

# 读取Excel行至数列
def row_to_list(sheet, row, column, count):
    resList = []
    for i in range(column, column + count):
        resList.append(str(sheet.cell(row=row, column=i).value).strip().lower())
    return resList

# some tests

wb = load_workbook(filename = "cpTest.xlsx")
wss = wb["Sheet1"]

strList = []

for i in range(2, 8):
    strList.append(','.join(row_to_list(wss, i, 4, 2)))

print (';'.join(strList))
