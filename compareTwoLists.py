#!/usr/bin/env python
# encoding: utf-8
from openpyxl import load_workbook
from openpyxl.styles import Font,colors

# 找第一个非空单元格
def find_row_count_at_column(sheet, row, column):
    count = 0
    for i in range(row, 3000):
        if sheet.cell(row=i, column=column).value is None:
            return count
        else:
            count += 1

# 数列写入Excel行
def list_to_row(data, sheet, row, column):
    for i in range(0, list(data).__len__()):
        sheet.cell(row=row, column=column + i).value = data[i]

# 读取Excel行至数列
def row_to_list(sheet, row, column, count):
    resList = []
    for i in range(column, column + count):
        resList.append(str(sheet.cell(row=row, column=i).value).strip())
    return resList

# 比较文件
FILE_NAME = 'q4.xlsx'
# 源表单名称
SRC_SHEET_NAME = 'Sheet1'
# 比较结果表单
COMPARE_RESULT = 'result'
# 核对结果表单位置
RESULT_SHEET_NUMBER = 1
# 首行号
FIRST_LINE = 2
# 总数据列数
COUNT_ALL = 4
# 关键字列数
COUNT_KEY = 3
# 是否循环比较
IS_LOOP_COMPARE = True
# 源列号
COLUMN_SRC = 1
# 目标列号 总数据列加1再加上1个分隔列
COLUMN_DES = COUNT_ALL + 1 + 1
# 列尾号
COLUMN_END = COUNT_ALL * 2 + 2

# 分隔符
DIV_CHAR = "&&"

# 加载文件
wb = load_workbook(filename = FILE_NAME)

# 若比较结果表单则先删除
try:
    ws = wb[COMPARE_RESULT]
    wb.remove(ws)
except KeyError as e:
    print ('No sheet test:' + e.__str__())

wss = wb[SRC_SHEET_NAME]

# 创建比较结果表单
wb.create_sheet(COMPARE_RESULT)
wsr = wb[COMPARE_RESULT]

# 写入标题行
for i in range(1, COLUMN_END):
    wsr.cell(row=1, column=i).value = wss.cell(row=1, column=i).value
    wsr.cell(row=1, column=i).font = Font(bold=True)

# 当前行号
currentRow = 2

# 获取源数据总数，目标数据总数
srcRowCount = find_row_count_at_column(wss, FIRST_LINE, COLUMN_SRC)
desRowCount = find_row_count_at_column(wss, FIRST_LINE, COLUMN_DES)

# 源数据，目标数据字典
srcDict = {}
desDict = {}

# 读取源数据进数据字典
for i in range(FIRST_LINE, FIRST_LINE + srcRowCount):
    key = DIV_CHAR.join(row_to_list(wss, i, COLUMN_SRC, COUNT_KEY))
    srcDict[key] = row_to_list(wss, i, COLUMN_SRC, COUNT_ALL)
# print (srcDict)

# 读取对比数据进数据字典
for i in range(FIRST_LINE, FIRST_LINE + desRowCount):
    key = DIV_CHAR.join(row_to_list(wss, i, COLUMN_DES, COUNT_KEY))
    desDict[key] = row_to_list(wss, i, COLUMN_DES, COUNT_ALL)
# print (desDict)

# 相同数据数
count_same = 0

# 循环源数据字典，在对比数据字典中找到相同的，则拷贝到新的结果页，同时在源数据字典和目标数据字典中都删除
for key in srcDict.copy():
    if key in desDict.keys():
        # 相同数据数加1
        count_same += 1
        # 相同数据写入目标页
        list_to_row(srcDict[key], wsr, currentRow, COLUMN_SRC)
        list_to_row(desDict[key], wsr, currentRow, COLUMN_DES)
        # 源数据、目标数据字典内去除相同数据
        del srcDict[key]
        del desDict[key]

        currentRow += 1

print ("There are " + str(count_same) + " same records")
print ("There are " + str(srcDict.__len__()) + " records can't find in des")
print ("There are " + str(desDict.__len__()) + " records can't find in src")

if IS_LOOP_COMPARE:
    for i in range(COUNT_KEY-1, 0, -1):
        count_same = 0
        # print (i)
        wsr.cell(row=currentRow, column=COLUMN_SRC).value = "以下数据前" + str(i) + "列" + "相同"
        wsr.cell(row=currentRow, column=COLUMN_SRC).font = Font(bold=True, color=colors.RED)
        currentRow += 1

        # 循环源数据字典，在对比数据字典中找到相同的，则拷贝到新的结果页，同时在源数据字典和目标数据字典中都删除
        # 重建源数据、目标数据字典，减少一最后一列值
        for key in srcDict.copy():
            keyList = key.split(DIV_CHAR)
            del keyList[-1]
            srcDict[DIV_CHAR.join(keyList)] = srcDict[key]
            del srcDict[key]

        for key in desDict.copy():
            keyList = key.split(DIV_CHAR)
            del keyList[-1]
            desDict[DIV_CHAR.join(keyList)] = desDict[key]
            del desDict[key]

        # 循环源数据字典，在对比数据字典中找到相同的，则拷贝到新的结果页，同时在源数据字典和目标数据字典中都删除
        for key in srcDict.copy():
            if key in desDict.keys():
                # 相同数据数加1
                count_same += 1
                # 相同数据写入目标页
                list_to_row(srcDict[key], wsr, currentRow, COLUMN_SRC)
                list_to_row(desDict[key], wsr, currentRow, COLUMN_DES)
                # 源数据、目标数据字典内去除相同数据
                del srcDict[key]
                del desDict[key]

                currentRow += 1

        print("Compare first " + str(i) + " columns: There are " + str(count_same) + " same records")
        print("Compare first " + str(i) + " columns: There are " + str(srcDict.__len__()) + " records can't find in des")
        print("Compare first " + str(i) + " columns: There are " + str(desDict.__len__()) + " records can't find in src")

wsr.cell(row=currentRow, column=COLUMN_SRC).value = "完全差异数据"
wsr.cell(row=currentRow, column=COLUMN_SRC).font = Font(bold=True, color=colors.RED)
currentRow += 1

for key in srcDict.keys():
    list_to_row(srcDict[key], wsr, currentRow, COLUMN_SRC)
    currentRow += 1

for key in desDict.keys():
    list_to_row(desDict[key], wsr, currentRow, COLUMN_DES)
    currentRow += 1

wb.active = RESULT_SHEET_NUMBER
wb.save(FILE_NAME)